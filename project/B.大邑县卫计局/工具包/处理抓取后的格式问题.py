# -*- coding=gbk -*-
import re
import openpyxl
from bs4 import BeautifulSoup
import sys


data = openpyxl.load_workbook(r'E:\Python\PyCharm\project\project1\卫计局2(20181025).xlsx')
# data2 = openpyxl.load_workbook(r'E:\Python\PyCharm\project\project1\123.xlsx')
active = data.active
# active2 = data2.active
table = data['Sheet1']
table2 = data['Sheet2']
rows = table.max_row
rows2 = table2.max_row

for row in range(2,rows+1) :
    # rowH = 'H%s'%(row)
    # print("正在处理第" +str(row)+"行数据")
    rowJ = 'J%s'%(row)  #  发布时间
    rowJ_va = table[rowJ].value
    rowK = 'K%s' % (row)  #
    rowK_va = table[rowK].value
    rowC = 'C%s' % (row)  #
    rowC_va = table[rowC].value
    key = 1
    rowC2  =  'C%s'%(row)
    rowD2  =  'D%s'%(row)
    rowE2  =  'E%s'%(row)
    rowF2  =  'F%s'%(row)
    rowG2  =  'G%s'%(row)
    rowH2  =  'H%s'%(row)
    rowI2  =  'I%s'%(row)
    rowJ2  =  'J%s'%(row)
    rowK2  =  'K%s'%(row)
    rowL2  =  'L%s'%(row)
    rowM2  =  'M%s'%(row)
    rowN2  =  'N%s'%(row)


    RSA1 = re.findall(r'<a.*?>',rowJ_va)
    RSA2 = re.findall(r'<img.*?>',rowJ_va)
    if RSA1!=[] or RSA2 !=[]:
        if RSA1!=[]:
            RSA3 = re.findall(r'<a.*?href="/datafolder/卫计局数据校正.*?"',rowJ_va)
            if RSA3!=[] :
                pass
            else:
                标题1  =  rowC_va
                第几行1 = row
                table2[rowC2] = 标题1
                table2[rowD2] = 第几行1
                table2[rowE2] = 'J'
        else:
            RSA4 = re.findall(r'<img.*?href="/datafolder/卫计局数据校正.*?"',rowJ_va)
            if RSA4 !=[]:
                pass
            else:
                table2[rowF2] = 标题1
                table2[rowG2] = 第几行1
                table2[rowH2] = 'J'

    if rowK_va !=None:
        RSA5 = re.findall(r'<a.*?>', rowK_va)
        RSA6 = re.findall(r'<img.*?>', rowK_va)

        if RSA5 != [] or RSA6 != []:
            if RSA5!=[]:
                RSA7 = re.findall(r'<a.*?href="/datafolder/卫计局数据校正.*?"',rowK_va)
                if RSA7!=[] :
                    pass
                else:
                    标题 =  rowC_va
                    第几行 = row
                    # print(str(标题) + str(第几行) + 'K')
                    table2[rowI2] = 标题1
                    table2[rowJ2] = 第几行1
                    table2[rowK2] = 'k'
        else:
                RSA8 = re.findall(r'<img.*?href="/datafolder/卫计局数据校正.*?"',rowK_va)
                if RSA8 !=[]:
                    pass
                else:
                    标题 = rowC_va
                    第几行 = row
                    table2[rowL2] = 标题1
                    table2[rowM2] = 第几行1
                    table2[rowN2] = 'k'

                    # table2[rowC] = rowC_va
    print("已处理第" +str(row)+"行数据")

# data2.save(r'E:\Python\PyCharm\project\project1\123.xlsx')
data.save(r'E:\Python\PyCharm\project\project1\卫计局2(20181025).xlsx')