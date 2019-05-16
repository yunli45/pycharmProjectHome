# -*- coding=gbk -*-


import re
import os,shutil
import openpyxl
import requests
from bs4 import BeautifulSoup

from ���ƾ�����У�� import �������ݿ�,Ԥ����ģ��,�������س���
from ���ƾ�����У��.�ж�urlǰ��ĵ㷵�������������ַ import returnSRC

header = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'}

def parePage (contenSrc):
    print(contenSrc)
    # contenSrc = 'http://www.nhfpc.gov.cn/zwgk/jdjd/201804/96ab7d7da43f45068d2afca782873d69.shtml'
    resposn = requests.get(contenSrc,headers=header)
    response = resposn.content.decode('utf-8', errors='ignore')
    soup = BeautifulSoup(response, 'lxml')

    soupCont1 = soup.find_all('div', attrs={'class': 'content'})  # z����
    soupCont2 = soup.find_all('div', attrs={'class': 'xw_box'})
    soupCont3= soup.find_all('div', attrs={'id': 'xw_box'})
    print(soupCont3)
    if soupCont3 !=[]:
        soupCont = soupCont3
    else:
        print("xw_box������ĸ�ʽ")
    # if soupCont1 != [] or soupCont2 != [] :
    #     if soupCont1 != []:
    #         soupCont = soupCont1
    #     elif soupCont2 != []:
    #         soupCont = soupCont2
    #     # print("soupCont2" + str(soupCont2))
    # else:
    #     print("����content��con���������ĸ�ʽ")
    print("soupCont"+str(soupCont))
    �����ӱ��ص�ַ = '/datafolder/���ƾ�����У��/��������ίԱ��/�ȵ���Ŀ/���'
    SavePath = r"F:\���ƾ�\���ƾָ���\��������ίԱ��\�ȵ���Ŀ\���\%s"
    indexUrl ='http://www.nhfpc.gov.cn/zwgk/jdjd/ejlist.shtml'
    content = Ԥ����ģ��.disposeOfData(indexUrl, contenSrc, str(soupCont[0]), SavePath, �����ӱ��ص�ַ)

    print(content)
    return content


def eachFileName(Filepath):
    # ��ȡ���ظ���������
    fileNameList = []
    for root, dirs, files in os.walk(Filepath, topdown=True):
        for name in files:
            file = os.path.join(root, name)
            file1 = file[file.rfind("\\")+1:]
            fileNameList.append(file1)
    return fileNameList

def seachFilePath(Filepath):
    # ��ȡ���ظ��������ĵ�ַ
    filePathList = []
    for root, dirs, files in os.walk(Filepath, topdown=True):
        for name in files:
            filePathList.append(os.path.join(root, name))
    return filePathList

def mkdir(path):
    # �жϴ�������·���Ƿ���ڣ������ھʹ���
    # ȥ����λ�ո�
    path = path.strip()

    # ȥ��β�� \ ����
    path = path.rstrip("\\")
    # �ж�·���Ƿ����
    # ����     True
    # ������   False
    isExists = os.path.exists(path)
    if not isExists:
        # ����������򴴽�Ŀ¼
        # ����Ŀ¼��������
        os.makedirs(path)
        print(path + "�����ɹ�")
    else:
        pass
        # print(path + "Ŀ¼�Ѵ���")

# ��ȡexcel���ļ�·���ͱ������е��ļ���·�����бȽϣ�·������ľͽ��ļ��Ƶ���ȷ��·���£����ز����ڵľʹ�ӡ����
def readExcelFile(Filepath,excelPath1,excelPath2):
    # ��ȡexcel���ļ�·���ͱ������е��ļ���·�����бȽϣ�·������ľͽ��ļ��Ƶ���ȷ��·���£����ز����ڵľʹ�ӡ����
    data = openpyxl.load_workbook(excelPath1)
    active = data.active
    table = data['Sheet1']
    rows = table.max_row

    data2 = openpyxl.load_workbook(excelPath2)
    active2 = data2.active
    table2 = data2['Sheet1']
    rows2 = table2.max_row
    # ��ȡ������������ļ���������·��
    FilePathList = seachFilePath(Filepath)
    # ��ȡ���ر�����������ļ�������
    FileNameList = eachFileName(Filepath)

    for row in range(2,rows+1) :
        # print("���ڶ�ȡ�ڣ�"+str(row)+"������")
        rowC = 'C%s'%(row) # ����
        rowC_Va = table[rowC].value
        # print(rowC_Va)

        rowJ = 'J%s'%(row) # ȫ��
        rowJ_Va = table[rowJ].value

        rowK = 'K%s' % (row)  # ȫ��
        rowK_Va = table[rowK].value

        rsFile = re.findall(r'<a.*?href="/(.*?)".*?>',str(rowJ_Va))
        # print(rsFile)

        # �Ƚϱ��ر���ĸ�����ȫ����������ĸ������ø����ڱ����Ƿ���ڣ������ھͻ��ڿ���̨��ӡ����Ӧ�ı���͸���λ��
        # if rsFile!=[]:
        #     rs = rsFile[0].replace("���ƾ�����У��",'')
        #     rs1 = r"F:\\���ƾ�����У��\\���ƾָ���"+rs.replace("/",r'\\')
        #     if rs1 in filList:
        #         pass
        #     else:
        #         print("��������ò���ڱ���û����Ӧ�ĸ�������鿴�¡���Ӧ�ı���Ϊ��"+str(rowC_Va))
        #         print("�������ݵĸ�������Ϊ��"+rs1)
        #         print("\n")
        # #
        #  �Ƚϱ������е��ļ��ĺ�excel���ļ�
        #     �ȱȽ�������·�����ڱȽ�����·��ƥ�䲻�ϵ����
                # ������·��ƥ�䲻��������£��Ƚ��ļ�����������ش�����Ӧ���ļ������Ǿ���λ�ô�����ˣ�����û����Ӧ���ļ����Ǿ���ȱʧ�ļ�

        # """"""
        if rsFile !=[]:
            rs = rsFile[0]
            if rs.find("@") != -1 or rs.find("www") != -1:
                pass
            else:
                # print(rs)
                # print(rowC_Va)
                # F:\\���ƾ�����У��\\���ƾָ���\\ʳƷ��ȫ��׼��������˾\\�����ļ�\\doc9458.doc
                # print("rs"+rs)
                ��������������ַƴ�� = rs .replace("���ƾ�����У��/",'')
                ��������������ַƴ�� = "F:\\���ƾ�����У��\\���ƾָ���\\"+��������������ַƴ��.replace("/","\\") # ���������� / �滻Ϊ \��ƴ��Ϊһ�������ı��ظ�����ַ
                ��������������ַƴ�� = ��������������ַƴ��.replace('\\',r'\\')  # ��ƴ�ӳ�����������ַ�� \ �滻Ϊ \\ ;F:\\���ƾ�����У��\\���ƾָ���\\ʳƷ��ȫ��׼��������˾\\�����ļ�\\doc9458.doc
                ��������������ַ = ��������������ַƴ��

                if ��������������ַ in FilePathList:
                    # print("���ļ��Ѵ���1")
                    pass
                else:
                    # ���ڿ�ʼ�Ƚ��ļ���
                    fileName = ��������������ַ[��������������ַ.rfind("\\")+1:]
                    # print("fileName"+fileName)
                    if fileName in FileNameList:
                        pass
                        # �ڱ��ش�����Ӧ���ļ���
                        newPath = ��������������ַ[:��������������ַ.rfind(r"\\")+2]
                        mkdir(newPath)
                        # print("newPath"+newPath)
                        if os.path.exists(newPath):
                            # print("���ļ��Ѵ���2")
                            pass
                        else:
                            for  path in FilePathList:
                                # ��������Ӧ���ļ���������·���ҳ���
                                if path.find(fileName) !=-1:
                                    oldFilepath = path
                                    shutil.move(oldFilepath,newPath)
                            break
                    else:
                        SQL = "select ����,�������ݵ���������url,��Դģ����ҳurl,�������ݵ�url,��Դ������ from ���������������� where ����='%s'" % (
                            str(rowC_Va).strip())
                        # SQL = "select ����,�������ݵ���������url,��Դģ����ҳurl,�������ݵ�url,��Դ������ from ���������������� where ����='%s' and ��Դ������='����˾>�����ļ�'" % (
                        #     str(rowC_Va).strip())
                        RS = �������ݿ�.getConnect(SQL)[2]
                        # print(RS)
                        keyId = 2
                        if RS != None:
                            # print(RS)
                            ���� = RS[0]
                            ������URL = RS[1]
                            indexUrl = RS[2]
                            url = RS[3]
                            ��Դ������ = RS[4]
                            print(��Դ������)
                            if ��Դ������ !="�л����񹲺͹�������������ίԱ��>�ȵ���Ŀ>���":
                                print("��������ȡ " + rowC_Va)
                                print(������URL)
                                # rowA2 = 'A%s' % (keyId)  # ����
                                # rowA2_Va = table2[rowA2].value
                                #
                                # rowB2 = 'B%s' % (keyId)  # ����
                                # rowB2_Va = table2[rowB2].value

                                # cont = parePage(������URL)
                                # table2[rowA2] = rowA2_Va
                                # table2[rowB2] = cont
                                # keyId +=1
                        else:
                            print("�ⲻ���ҵ�  "+str(rowC_Va))
    data2.save(excelPath2)


    print(keyId)

# �Ƚ�excel�г����ӵ��ļ����ͱ����ļ�
def compareExcelAndFile(Filepath,excelPath):
    data = openpyxl.load_workbook(excelPath)
    active = data.active
    table = data['Sheet1']
    rows = table.max_row
    keyId = 1

    # ��ȡ������������ļ���������·��
    FilePathList = seachFilePath(Filepath)
    # ��ȡ���ر�����������ļ�������
    FileNameList = eachFileName(Filepath)
    # titleList =['����˲����ι���취���޶�����2013��3��24����ʵʩ', '��������������ί��������ũ�����ض�ͯ�����ذ�������֪ͨ�����', '�����ڽ�һ����ǿ����ҩ���ٴ�Ӧ�ù������ϸ����ҩ��֪ͨ�����', '��ҽʦִҵע�����취�����', '��������������ί�칫�����������°��޳���Ѫ֤��֪ͨ���Ľ��', '�������ٴ�������Ŀ�����й������֪ͨ�����', 'һͼ��������������+ҽ�ƽ������������', '�����ڼ�ǿ����ҽԺ���Ľ��蹤������������߽��', '��ʳƷ��ȫ���ұ�׼ ֲ���͡���׼�Ƚ������', '2018���סԺҽʦ�淶����ѵ������ȫ��ҽ����ѵ�������ҵ���˹������߽��', 'һͼ���������ڴٽ���������+ҽ�ƽ�������չ�������', '��������������ί�칫������ӡ����Ⱦ�Լ�����ظ��廯ҽѧ���Ӽ�⼼��ָ�Ϻ͸��廯ҽѧ���΢���л���оƬ�����淶��֪ͨ�����', '����2017�����ҽҩ������ҽ�Ʒ����в���֮��ר��������Ҫ��Ľ��', '����ҽ�úĲ�ר�����λ�����Ľ��', 'һͼ����|��ȱҩ���ϣ�����Ҫ����Щ��', '��ҽ��X�������Ʒ������Ҫ�󡷽��', '2017��סԺҽʦ�淶����ѵ��ҵ���˹������߽��', '��ʳƷ��ȫ���ұ�׼ ʳƷ�����������������GB 2761-2017������ʳƷ��ȫ���ұ�׼ ʳƷ����Ⱦ��������(GB 2762-2017)���', 'ͼ��ʳƷ��ȫ��׼����������ʮ���塱�滮��2016-2020�꣩', '����ʮ���塱ȫ����˲����ι滮��ͼ��', '����ҽѧӰ��������ĵȶ�������ҽ�ƻ���������׼�͹���淶���', '�����ڼ�ǿ�����������ָ�������һͼ����', 'ͼ������Ժӡ������ʮ���塱�����뽡���滮��', '���ڡ��漰�˵�����ҽѧ�о��������취���Ľ��', 'һͼ�����������й�2030���滮��Ҫ', '����������Ͷ������ױ���Ժ�����׼��ʵʩϸ��2016��棩', '�����ڴٽ��͹淶����ҽ�ƴ�����Ӧ�÷�չ��ָ����������', '����ҩƷ�۸�̸���й����˵��', 'ͼ�⣺��ҽҩ��չս�Թ滮��Ҫ��2016��2030�꣩', '�����ڸ��׽������������׼��������淶�������ָ�������������ļ����', 'ͼ�⣺��һ������ҽ�Ʒ����ж��ƻ�', 'ͼ�⣺�й�����Ӧ������', '������Դ��ᱣ�ϲ��͹�����������ί����ӡ�������ڽ�һ���ĸ����ƻ�������רҵ������Աְ����������ָ�������', '����������Ԥ������������Ҫ�󡷽��', '������ӡ����Ⱦ����Ϣ�������淶��2015��棩��֪ͨ���ļ����', 'ͼ�⣺�ҹ��ƻ�������������', 'ͼ�⣺��ʮ���塱������ս��', '���й���֢���������ж��ƻ���2015-2017�꣩�����', 'ͼ�⣺�й�����Ӫ�������Բ�״�����棨2015�꣩', 'ͼ�⣺ȫ���������������滮��2015��2020�꣩', 'ͼ�⣺����Ժ�칫�����ڳ��й���ҽԺ�ۺϸĸ��Ե��ָ�����', 'ͼ�⣺�ҽҩ�������Ƹĸ�2015���ص㹤������', 'ͼ�⣺�ҽҩ�������Ƹĸ�2014�깤���ܽ�', 'ͼ�⣺����ȫ���ƿ��ؼ�����ҽԺ�ۺϸĸ��ʵʩ���', 'ͼ�⣺���Ρ����ǡ�ר���ж�ʵʩ����', 'ͼ�⣺����Ժ�칫��ת���������Ȳ��Ź��ڽ�һ������ҽ�ƾ����ƶ�ȫ�濪չ���ش󼲲�ҽ�ƾ������������֪ͨ', '���ٴ���������������Ŀ�ο������5���֣�Ѫ�����ء���������4���Ƽ���������ҵ��׼���', 'ͼ�⣺ְҵ����������취(���������ͼƻ�����ίԱ�����5��)', 'ͼ�⣺���������Ƽ���������2014���չ��2015�깤������', 'ͼ�⣺�й�����Ԥ�����ƹ�����չ��2015�꣩', 'ͼ�⣺�в��������������', 'ͼ�⣺�������ר���ж���������', 'ͼ�⣺ȫ�����׽����������³ɼ��������������ص�', 'ͼ�⣺��ʶ���һ���������������', '������Ʒ������ȫ���۹涨��������Ʒ�����ල�����淶���', '�׾Ʋ�Ʒ���ܻ�����������������', '���˿ڽ�����Ϣ����취�����У�������������壩���', '������������ί���������ల�ͼ�ּƻ�����������������ʵʩ�����������ߴ������', '������������ί֪ͨ�Ͻ�ҽ�ƻ���������Ա��������ĸ�����Ʒ', '��������ί��Ӧ��ḧ�������չ����й�����', '������������ί������չ���������ලִ��ר�����', '������������ί�͡��漰�����ҽѧ��ѧ�����о�����취�������������', '������������ί����ʳƷ��ȫ���ұ�׼��ʳƷ����ͨ�������淶����GB14881-2013��', '������������ί���������涨����', '��������ί�淶��������������Ƽ���������', '������������������Ҫ�㼰��ؽ���', '��ְҵ��������������취�����', '��ʳƷ����Ⱦ����������GB2762-2012���ʴ�', '���ڡ�ְҵ�������Ŀ¼�������Ĵ���', '��������6��������ӡ������ֵҽ�úĲļ��вɹ������淶�����У���', '����˲����ι���취���޶���������壩����д˵��', '���ڽ�ֹ��������λ�ɹ��������ʹ���������ε�˵��', '���ڡ��й�����Σ���������桷�Ĵ���', '���ж����׶�԰�������������淶���޶�˵��', '���ڡ��ж����׶�԰�������������淶���Ĵ���', '�������ƽ�����ũ�����ҽ��֧����ʽ�ĸ﹤����ָ��������ʴ�', '������ҩ���ٴ�Ӧ�ù���취���й��������', '��������ҵ���ջ�����������ũ�����ҽ�ƾ�������ָ��������ʴ�', '��������׼�ϸ���ɫ�ص�9��ʳƷ��Ӽ�', '���������������301��ʳƷ��װ��������Ӽ�����', 'ʳ����������������֪ʶ�ʴ�', '������������ʳƷӪ��ǿ����ʹ�ñ�׼��(GB 14880-2012��', 'ȫ���������������ű�������������й�ȫ������˲���ű����Ŀ�������ڹ����ٿ�', '�������ල�ֹ��ڿ�չ�����ලר����鹤����֪ͨ', '�������칫�����ڿ�չҽ�÷������������Ե㹤����֪ͨ', '2009������ͳ�ƹ������ݽ��', 'Ԥ������֪ʶ�ȵ�������ʲ��ϣ�����', '���������ҩƷ��ѵ��Ŀ������ͼ��Ԣ��ڹ��', '�йظ����˾ͽ������һ���ҩ���ƶȴ���', '�������칫���������󡶵��Ӳ��������ܹ������ݱ�׼����������壩�������֪ͨ', '�ҽҩ�������Ƹĸﲿ��Э������С�鸺���˸����˾�ҽ�ķ���������������', '��������������ӡ����ҽѧ�����ٴ�ʵ���������й涨��', '������ӡ�������ҽ�����˰취��', '���������������涨����', '���봨������������Ԥ�����ƶԿ�֧Ԯ���������ӡ��', '������ӡ��������������֯����ϸ������֯��������ֲָ��ԭ�򣨲ݰ�����', '������ӡ����������Ѫ�ȵ�6�ִ�Ⱦ��Ԥ������ָ�Ϻ��ٴ����Ʒ���', '�����������ҷ���ί����2008��2010�������ӷ����Ͷϲ�ɨ�������ù滮', '�ɹ�Ѫ������ҵ��ԱҪʵ�и�λ��ѵ�����ƶ�', '����������Ԥ�����ƾָ����˺����з���ר�ҽ����й�������ר��̸�����Ը�ð�ķ���', '���Ծ��񼲲����������Ŀ��ʼʵʩ', '����������Ԥ�����ƾֺ��й�����Ԥ���������ķ����й����Բ�����', '����������2005�ꡰ����������������ʹ���߾��񽡿�����ʹ����ѡ���', '2005��ȫ��ʮ���������Ž��ս���', '�ܾ�������������������׼����?�ٽ�������ҵ��չ����������ͨ������22��']
#

    # for row in range(2200, rows+1):
    # for row in range(2500, rows+1):
    for row in range(1, rows+1):

        print("���ڶ�ȡ�ڣ�"+str(row)+"������")
        rowC = 'C%s' % (row)  # ����
        rowC_Va = table[rowC].value
        # print(rowC_Va)

        rowJ = 'J%s' % (row)  # ȫ��
        rowJ_Va = table[rowJ].value

        rowK = 'K%s' % (row)  # ȫ��
        rowK_Va = table[rowK].value
        print(1)
        rsA = re.findall(r'<a href="(.*?)"(.*?)>(.*?)</a>', rowJ_Va)
        print(1.1)
        if rsA !=[] :
            print(2)
            print("���ǵڼ���  "+ str(row))
            for A in rsA:
                print(row)
                print(rowC_Va)
                print(A)
                # �滻���������
                if str(A[0]).find("shtml") != -1 or str(A[0]) == str(A[2]) or str(A[0]).find("htm") != -1 or str(
                        A[0]).find(
                        "html") != -1 or "http://"+str(A[2]) == str(A[0]) :
                    rowJ_Va = rowJ_Va.replace(r'<a href="%s"%s>%s</a>' % (str(A[0]), str(A[1]), str(A[2])),
                                              "%s" % (str(A[2])))
            table[rowJ] = rowJ_Va
            print("�Ѿ������˸�������" + str(rowC_Va))
            print(rowJ_Va)
        else:
            print(3)
            pass

    data.save(excelPath)





Filepath = r'F:\���ƾ�\���ƾָ���'
# fileList = seachFile(Filepath)
# print(fileList)
excelPath1 = 'E:\Python\PyCharm\project\project1\���ƾ�\(20181031).xlsx'
excelPath2 = 'E:\Python\PyCharm\project\project1\���ƾ�\(110).xlsx'
readExcelFile(Filepath,excelPath1,excelPath2)
# parePage('')

# ��ȡ������������ļ���������·��
# FilePathList = seachFilePath(Filepath)
# ��ȡ���ر�����������ļ�������
# FileNameList = eachFileName(Filepath)