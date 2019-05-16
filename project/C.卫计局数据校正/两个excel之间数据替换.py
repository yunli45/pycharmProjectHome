import re
import openpyxl



# 表一为重新抓取的数据，二为源数据表
def readExcel(excelPath1,excelPath2):
    data1 = openpyxl.load_workbook(excelPath1)
    active1 = data1.active
    table1 = data1['Sheet1']
    rows1 = table1.max_row

    data2 = openpyxl.load_workbook(excelPath2)
    active2 = data2.active
    table2 = data2['Sheet1']
    rows2 = table2.max_row


    listTitle = []
    for row1 in range(2,rows1+1):
        #     print("正在读取第："+str(row)+"的数据")
        标题1 = 'A%s' % (row1)  # 标题
        标题1_Va = table1[标题1].value
        # print(rowC_Va)

        全文1 = 'C%s' % (row1)  # 数据
        全文1_Va = table1[全文1].value

        全文2 = 'D%s' % (row1)  # 数据
        全文2_Va = table1[全文2].value

        序号1 = 'B%s' % (row1)  # 序号
        序号1_Va = table1[序号1].value

        for row2 in range(2,rows2+1):

            标题2 = 'C%s' % (row2)  # 标题
            标题2_Va = table2[标题2].value

            全文3 = 'J%s' % (row2)  # 数据
            全文3_Va = table2[全文3].value

            全文4 = 'K%s' % (row2)  # 数据
            全文4_Va = table2[全文4].value

            if 序号1_Va == str(row2) and 标题1_Va == 标题2_Va:

                # print("正在处理第 " + str(row2) + "的数据")
                if 全文4_Va!=None:
                    table2[全文3] = 全文1_Va
                    table2[全文4] = 全文2_Va
                    print("已处理的标题 "+ 标题2)
                elif 全文4_Va ==None  and 全文3_Va !=None:
                    table2[全文3] = 全文1_Va
                    print("已处理的标题 " + 标题2)
                else:
                    print("特殊情况  " +标题2 )
            else:
                print("2")
    data2.save(excelPath2)

excelPath1 = 'E:\Python\PyCharm\project\project1\卫计局\根据标题修改文本.xlsx'
excelPath2 = 'E:\Python\PyCharm\project\project1\卫计局\卫计局2(20181025).xlsx'
readExcel(excelPath1,excelPath2)