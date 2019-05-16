# -*- coding: utf-8 -*-
"""
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
              ┏┓      ┏┓
            ┏┛┻━━━┛┻┓
            ┃      ☃      ┃
            ┃  ┳┛  ┗┳  ┃
            ┃      ┻      ┃
            ┗━┓      ┏━┛
                ┃      ┗━━━┓
                ┃  神兽保佑    ┣┓
                ┃　永无BUG！   ┏┛
                ┗┓┓┏━┳┓┏┛
                  ┃┫┫  ┃┫┫
                  ┗┻┛  ┗┻┛
"""
import openpyxl


# 读取excel文件返回该文件所有数据和最大的行数，保存数据的时候记得使用返回的第一个参数 【data.save(excelFilePath)】
def open_excel_file(excel_file_path):
    data = openpyxl.load_workbook(excel_file_path)
    active = data.active
    table = data['Sheet1']
    rows = table.max_row

    return data, table, rows

def witer_excel_flie(excel_file_path):
     openpyxl.Workbook()

