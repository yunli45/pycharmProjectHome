# coding = utf-8
import openpyxl
import os,shutil
import re
from 工具包 import 链接数据库


# 读取excel文件返回该文件所有数据和最大的行数，保存数据的时候记得使用返回的第一个参数 【data.save(excelFilePath)】
def open_excel_file(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)# 打开一个已有的workbook
    ws = wb['Sheet1']
    rows = ws.max_row

    return wb, ws, rows

# 删除本地sheet，并新建两个sheet，如果存在相应的表格会重新新建并覆盖之前的表格
def del_create(adj_excel_file_path):
    """
        将本地用于保存附件相关情况的表格先删除所有的sheet，并新建两个sheet：附件情况，附件本地错误情况
    :param excel_file_path: 本地保存附件情况的excel所在全路径
    :return:
    """
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # wb.save('s.xlsx')

    # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb = openpyxl.Workbook()
    # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
    ws = wb.active
    wb.create_sheet('附件具体情况', 0)
    wb.create_sheet('附件大小为0的情况', 1)
    wb.save('{0}'.format(adj_excel_file_path))


# 获取本地所有附件的名字
def each_file_name(files_path):
    """
    :param files_path:  本地附件所在的总目录，以省份为总目录；
    :return: 返回一个集合
    """
    file_name_list = []
    for root, dirs, files in os.walk(files_path, topdown=True):
        for name in files:
            file = os.path.join(root, name)
            file1 = file[file.rfind("\\")+1:]
            file_name_list.append(file1)

    return file_name_list


# 获取本所有附件完整的地址
def each_file_path(files_path):
    """
    :param files_path:  本地附件所在的总目录，以省份为总目录；注意在与全文中的地址进行判断的时候记得在本地路径前面加上datafolder
    :return:
    """
    file_path_list = []
    for root, dirs, files in os.walk(files_path, topdown=True):
        for name in files:
            file_path_list.append(os.path.join(root, name))

    return file_path_list


# 判断传进来的路径是否存在，不存在就创建
def mk_dir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)
        print(path + "创建成功")
    else:
        print(path + "目录已存在")
        pass


# 复制文件到另一个目录下
def copyFileto(sourceDir,  targetDir):
    """
    :param sourceDir: 资源文件的完整路径，eg:E:\行政案例附件\datafolder\历史\河北省\廊坊市\所有行业\1.pdf
    :param targetDir: 复制后目的完整路径（也可以不是完整的路径，需要文件所在的目录：E:\行政案例附件\datafolder），eg:E:\行政案例附件\datafolder\1.pdf
    :return:
    """
    shutil.copy(sourceDir,  targetDir)


# 字节bytes转化kb\m\g
def formatSize(bytes):
    try:
        bytes = float(bytes)
        kb = bytes / 1024
    except:
        print("传入的字节格式不对")
        return "Error"

    if kb >= 1024:
        M = kb / 1024
        if M >= 1024:
            G = M / 1024
            return "%fG" % (G)
        else:
            return "%fM" % (M)
    else:
        return "%fkb" % (kb)


# 查看文件的大小:传入文件的全路径
def getfileSize(file_path):
    try:
        size = os.path.getsize(file_path)
        # print(formatSize(size))
        return formatSize(size)
    except Exception as err:
        raise Exception (err)


# 本地附件大小等于 0kb ,写入到本地一个文件中
def sizeOrrWiter( label, href, adj_local_Path, title, source_url, adj_excel_file_path, n):
    """
    :param txtFile:  本地附件等于 0kb 需要写入的文件路径
    :param label: 根据这个标记分类，判断本地附件的情况;
        0：本地附件存在，文中超链接也修改了，但是本地附件大小等于 0kb;
        1:
    :param href: 这个附件在文中的超链接
    :param adj_local_Path: 这个附件在本地保存的位置
    :param atitle: 这条数据的标题
    :param source_url: 这条数据的来源url
    :return: 不需要返回
    """
    wb = openpyxl.load_workbook(adj_excel_file_path)  # 打开一个已有的workbook
    ws = wb["附件具体情况"]
    ws['A1'].value = "标题"
    ws['B1'].value = "这条数据完整的访问地址"
    ws['C1'].value = "附件存在链接修改，附件在本地大小为 0kb"                                        # label = 0
    ws['D1'].value = "附件存在，链接修改本地位置错误（已复制到文中的地址），附件在本地大小为 0kb"        # label = 1
    ws['E1'].value = "附件不在链接修改了"                                                           # label = 2
    ws['F1'].value = "附件存在链接没改。本地附件为 0kb"                                              # label = 3-1
    ws['G1'].value = "附件存在链接没改"                                                             # label = 3-2
    ws['H1'].value = "附件不在链接没改"                                                             # label = 4

    # wb.save(adj_excel_file_path)


    if n == '':
        pass
    else:

        try:
            size = getfileSize(r'{0}'.format(adj_local_Path))
        except Exception:
            pass
        if label == 0:
            if size == '0.000000kb':
                # 链接修改附件存在且位置正确,附件在本地的大小是 0kb
                print("链接修改附件存在且位置正确,附件在本地的大小是 0kb")
                print(title)
                print(label, href, adj_local_Path, title, source_url, adj_excel_file_path, n)
                ws['C{0}'.format(n)] =  "文中的超链接：{0}, 本地附件：{1}".format(href, adj_local_Path)
                ws['A{0}'.format(n)].value = title
                ws['B{0}'.format(n)].value = source_url

                wb.save(adj_excel_file_path)
            else:
                pass
        elif label == 1:
            if size == '0.000000kb':
                print("本地保存位置有问题，现在已经将本地文件复制到文中的地址下了，附件在本地的大小是 0kb，")
                ws['D{0}'.format(n)] = "文中的超链接：{0}, 本地附件：{1}".format(href, adj_local_Path)
                ws['A{0}'.format(n)].value = title
                ws['B{0}'.format(n)].value = source_url

                wb.save(adj_excel_file_path)
            else:
                pass
        elif label == 2:
            print("附件不在链接修改了")
            ws['E{0}'.format(n)] = "文中的超链接：{0}, 本地附件：{1}".format(href, adj_local_Path)
            ws['A{0}'.format(n)].value = title
            ws['B{0}'.format(n)].value = source_url

            wb.save(adj_excel_file_path)
        elif label == 3:
            if size == '0.000000kb':
                print("附件存在链接没改。本地附件为 0kb")
                ws['F{0}'.format(n)] = "文中的超链接：{0}, 本地附件：{1}".format(href, adj_local_Path)
                ws['A{0}'.format(n)].value = title
                ws['B{0}'.format(n)].value = source_url

            else:
                print("附件存在链接没改。")
                ws['G{0}'.format(n)] = "文中的超链接：{0}, 本地附件：{1}".format(href, adj_local_Path)
                ws['A{0}'.format(n)].value = title
                ws['B{0}'.format(n)].value = source_url

                wb.save(adj_excel_file_path)

        elif label == 4:
            print("附件不在链接没改")
            ws['H{0}'.format(n)] = "文中的超链接：{0}, 本地附件：{1}".format(href, adj_local_Path)
            ws['A{0}'.format(n)].value = title
            ws['B{0}'.format(n)].value = source_url
            wb.save(adj_excel_file_path)

        else:
            raise Exception("传入的标记参数有误，只能是 0-4 !!!!")


# 检查附件、文中链接情况、处理
def check_adj(title,adj_file_path, cont, source_url, adj_orr_excel_path, n):
    """
     检查文章的附件地址修改了没，以及本地附件是否存在，共计4种情况：
        1、地址修改了 附件存在(还存在一种地址修改了，但是附件位置保存错了，直接按照文中的地址在本地创建相应的文件，并把附件复制到这个目录下。就不用返回了)  2、地址修改了 附件不在  3、地址没改  附件存在  4、地址没改 附件不在

    :param title: 这条数据的标题
    :param adj_file_path: 本地附件所在的总目录（eg E:\行政案例附件\datafolder\历史\河北省）
    :param cont: 这条数据的全文
    :param adj_excel_file_path: 本地附件检查情况保存的文件名
    :return: 2、地址修改了 附件不在  3、地址没改  附件存在  4、地址没改 附件不在 着三种情况到一个文本中
    """
    del_create(adj_orr_excel_path)  # 创建表格
    all_file_name_list = each_file_name(adj_file_path)  # 找到本地所有附件的名字
    all_file_path_list = each_file_path(adj_file_path)  # 找到本地所有附件的路径



    # 先查找所有的a标签，查看本地的附件
    # all_a = re.findall("""<a.*?href="(.*?)".*?>(.*?)</a>""", str(cont), flags=re.I)
    file_path_1 = adj_file_path[:adj_file_path.rfind("\\")]   # 本地附件总目录上一级
    all_a = re.findall(r'<a.*?href=".*?".*?>.*?</a>', str(cont), flags=re.I)
    n = n
    if all_a !=[]:
        for a in all_a:
            # 循环所有的a标签,再遍历每一个a标签，在每一个a标签中去匹配是否是附件形式 ，不是就直接替换掉，是的话在进行下一步：匹配附件a标签中冲链接和文本能容
            annex_a = re.findall(r'<a.*?href=".*?(pdf|docx|doc|xlsx|xls|rar|zip|jpeg|jpg|png|gif|txt|7z|gz)".*?>.*?</a>', a)
            if annex_a:
                print("这个a标签是附件")
                # 假如是附件的形式，使用all_a结果集中每个元素的第一个值（超链接）和第二个值（文本内容）反向匹配这个标签的全部内容，并生成新的本地超链接a标签
                rs_a = re.findall('<a.*?href="(.*?)".*?>(.*?)</a>', a)
                a_hyper = rs_a[0][0]  # a标签的超链接
                # a_text = rs_a[0][1]   # a标签的文本内容
                a_annex_name = a_hyper[a_hyper.rfind("/") + 1:]  # 附件名字

                # 匹配本地附件是否存在
                # 超链接地址已经做了修改是属于我们服务器上面的地址，文中的超链接修改过了，开始检查本地附件
                if a_hyper.find("/datafolder/") != -1:
                    a_hyper_local = file_path_1 + a_hyper.replace("/datafolder", '').replace("/", "\\")    # 替换、拼接文中超链接为本地的附件;也作为本地附件保存错误后需要复制到的正确地址

                    if a_hyper_local in all_file_path_list:
                        print("附件存在，链接修改，开始检查本地附件")
                        # 这个附件文中的超链接地址已经修改了，且本地的附件也是存在的,现在开始检查本地附件的大小，如果大小等于 0.000000kb 那么这个附件就不正确
                        sizeOrrWiter(0, a, a_hyper_local, title, source_url, adj_orr_excel_path, n)
                        pass
                    else:
                        # 这个附件文中的地址修改过了，本地的附件还不知道存在不。开始本地是否包含这个附件、以及本地的存储地址，本地不存在就是缺失，存在的话就是本地放错了位置，以文中的地址为准# # # # # # # # # # # # # # # # # "+str(title)+"  "+str(a_hyper)

                        if a_annex_name in all_file_name_list:
                            for local_file_path in all_file_path_list:
                                #  找到了这个文件名在本地的路径，说明文中的地址修改了，本地存在相应的附件开始将这个附件复制到文中地址的路径下
                                if local_file_path.find(a_annex_name) != -1:
                                    a_hyper_local_1 = a_hyper_local[:a_hyper_local.rfind("\\")]   # 创建附件所在的上一级目录（文中的地址）
                                    mk_dir(a_hyper_local_1)
                                    copyFileto(local_file_path, a_hyper_local)
                                    sizeOrrWiter(1, a, a_hyper_local, title, source_url, adj_orr_excel_path, n)
                                    break
                        else:
                            # 本地附件不存在，附件缺失且文中的 地址修改了的
                            # 附件不在链接修改了['{0}'.format(title)]= a
                            sizeOrrWiter(2, a, file_path_1, title, source_url, adj_orr_excel_path, n)

                else:
                    # 文中的超链接没有修改，开始检查本地附件是否存在，附件名字存在说明：附件存在，链接没改；不存在：附件不在，链接也没改
                    if a_annex_name in all_file_name_list:
                        # 附件在本地是存在的但是文中的地址是没有修改的,该附件在本地是存在的，只是超链接的地址没有修改，请手动修改,后期改为自动
                        for adj_name in all_file_path_list:
                            if adj_name.find(a_annex_name)!=-1:
                                # 本地的全路径名
                                local_path = adj_name
                                sizeOrrWiter(3, a, local_path, title, source_url, adj_orr_excel_path, n)
                                break
                    else:
                        # 该附件不存在，并且超链接的地址也没有修改，请重新下载该数据
                        sizeOrrWiter(4, a, '', title, source_url, adj_orr_excel_path, n)
            else:
                print("这个a标签是跳转的情况，现在把它的跳转去掉")
                a_text = re.findall(r'<a.*?href=".*?".*?>(.*?)</a>', a)
                before_a = a
                new_a1 = '{0}'.format(a_text[0])
                cont = cont.replace(before_a, new_a1)

    all_img = re.findall('<img.*?src=".*?".*?>', cont)
    if all_img != []:
        for img in all_img:
            picture = re.findall('<img.*?src=".*?(jpeg|jpg|png|gif)".*?>', img)
            if picture !=[] :
                rs_img = re.findall('<img.*?src="(.*?)".*?>', img)
                img_hyper = rs_img[0]   # 图片文中地址
                img_annex_name = img_hyper[img_hyper.rfind("/")+1 :]
                if img_hyper.find("/datafolder/")!= -1:
                    # 文中的链接是修改了的，开始检查附件存在与否以及与文中的地址是否相同
                    img_hyper_local = file_path_1 + img_hyper.replace("/datafolder", '').replace("/", "\\")    # 替换、拼接文中超链接为本地的附件;也作为本地附件保存错误后需要复制到的正确地址
                    if img_hyper_local in all_file_path_list:
                        # 文中的附件地址修改了，且本地也存在正确的位置, 开始检查这个附件的大小，等于 0kb 就是有问题
                        sizeOrrWiter(0, img, img_hyper_local, title, source_url, adj_orr_excel_path, n)
                        pass
                    else:
                        # 文中的地址修改了，本地不存在与文中相同的路径下的附件，开始检查本地是否存在相应的附件名
                        if img_annex_name in all_file_name_list:
                            # 本地存在相应的附件名但是本地的地址是错的，现在将附件创建文中的地址的并复制到
                            for local_img_path in all_file_path_list:
                                if local_img_path.find(img_annex_name)!=-1:
                                    img_hyper_local_1 = img_hyper_local[img_hyper_local.rfind("\\")]
                                    mk_dir(img_hyper_local_1)
                                    copyFileto(local_img_path, img_hyper_local_1)
                                    sizeOrrWiter(1, img, img_hyper_local, title, source_url, adj_orr_excel_path, n)
                                    break
                        else:
                            # 本地附件不存在，附件缺失且文中的 地址修改了的
                            sizeOrrWiter(2, img, '', title, source_url, adj_orr_excel_path, n)

                else:
                    # 文中没有修改地址，开始检查附件是否存在
                    if img_annex_name in all_file_name_list:
                        # 本地存在相应的附件，文中没有修改地址
                        for img_name_local in all_file_path_list:
                            if img_name_local.find(img_annex_name) !=-1:
                                sizeOrrWiter(3, img, img_name_local, title, source_url, adj_orr_excel_path, n)
                                break
                    else:
                        # 该附件不存在，并且超链接的地址也没有修改，请重新下载该数据
                        sizeOrrWiter(4, img, file_path_1, title, source_url, adj_orr_excel_path, n)
            else:
                # 这个img标签不是附件，是跳转，去除跳转
                old_img = img
                new_img1 = ''
                cont = cont.replace(old_img, new_img1)

    return cont

# 处理标签格式，不包括 a 、img
def processing_text(cont):
    """
    :param cont: 传一个文章的参数就行了，主要用于处理文中的各种标签问题，注意a img标签已经在前面处理好了。可以不用处理
    :return: cont：处理后返回处理后的文章，用于数据库存储
    """
    content = str(cont).replace("　", '').replace("	", '').replace(" ", '')
    # content = re.sub('<ul.*?>.*?</ul>',  '',  content, flags=re.S | re.I)
    content = re.sub(r'\f', '/', content)
    content = re.sub(r'\\', '/', content)
    content = re.sub(r'<font.*?>', '', content, flags=re.I).replace('</font>', '').replace('</FONT>', '')
    content = re.sub(r'<b.*?>', '', content, flags=re.I).replace('</b>', '').replace('</B>', '')
    content = re.sub(r'<span.*?>', '', content, flags=re.I).replace('</span>', '')
    content = re.sub(r'<col.*?>', '<col>', content, flags=re.I)

    """
       找到所有的p标签，在所有p标签的集合中再匹配出<p.*?text-align>、<p.*?align>、<p.*?style=".*?(text-align|align)
       最终替换掉原文中的p标签
    """
    all_p = re.findall('<p.*?>', content, flags=re.I)
    if all_p:
        for ids, p in enumerate(all_p):
            p_format_1 = re.findall('<p.*?(text-align|align)="(.*?)".*?>', p, re.I)
            p_format_2 = re.findall('<p.*?style=".*?(text-align|align):.*?".*?>', p, re.I)
            if p_format_1:
                content = content.replace(p, '<p align="%s">' % (p_format_1[0][1]))
            elif p_format_2:
                for pp in p_format_2:
                    p_format_2_1 = re.findall('<p.*?style=".*?(text-align|align):(right|left|center).*?".*?>', p, re.I)
                    if p_format_2_1:
                        content = content.replace(p, '<p align="%s">' % (p_format_2_1[0][1]))
                    else:
                        content = content.replace(p, '<p>')
            else:
                content = content.replace(p, '<p>')

    all_div = re.findall('<div.*?>', content, flags=re.I)
    if all_div:
        for ids, div in enumerate(all_div):
            div_format_1 = re.findall('<div.*?(text-align|align)="(.*?)".*?>', div, flags=re.I)
            div_format_2 = re.findall('<div.*?style=".*?(text-align|align):.*?".*?>', div, flags=re.I)
            if div_format_1:
                content = content.replace(div, '<div align="%s">' % (div_format_1[0][1]))
            elif div_format_2:
                for divdiv in div_format_2:
                    div_format_2_1 = re.findall('<div.*?style=".*?(text-align|align):(right|left|center).*?".*?>', div,
                                                flags=re.I)
                    if div_format_2_1:
                        content = content.replace(div, '<div align="%s">' % (div_format_2_1[0][1]))
                    else:
                        content = content.replace(div, '<div>')
            else:
                content = content.replace(div, '<div>')

    all_span = re.findall('<span.*?>', content, flags=re.I)
    if all_span:
        for ids, span1 in enumerate(all_span):
            span_format_1 = re.findall('<span.*?(text-align|align)="(.*?)".*?>', span1, flags=re.I)
            span_format_2 = re.findall('<span.*?style=".*?(text-align|align):.*?".*?>', span1, flags=re.I)
            if span_format_1:
                content = content.replace(span1, '<span align="%s">' % (span_format_1[0][1]))
            elif span_format_2:
                for span2 in span_format_2:
                    span_format_2_1 = re.findall('<span.*?style=".*?(text-align|align):(right|left|center).*?".*?>',
                                                 span1,
                                                 flags=re.I)
                    if span_format_2_1:
                        content = content.replace(span1, '<span align="%s">' % (span_format_2_1[0][1]))
                    else:
                        # print(p)
                        content = content.replace(span1, '<span>')
            else:
                content = content.replace(span1, '<span>')

    # 现在将table 设置为居中，显示边框。tr、td的属性全部不需要
    content = re.sub('<table.*?>', '<table>', content)
    content = re.sub('<tr.*?>', '<tr>', content)
    content = re.sub('<td.*?>', '<td>', content)

    table_existence = re.findall('<table', content, flags=re.I)
    tr_existence = re.findall('<tr', content, flags=re.I)
    if table_existence and tr_existence:
        content = re.sub('<table.*?>', '<table border="1" cellspacing="0" align="center">', content)
        if re.findall('</table>', content):
            pass
        else:
            content = content + '</table>'
    elif table_existence == [] and tr_existence:
        content = '<table border="1" cellspacing="0">\n' + content + '\n</table>'
    elif table_existence and tr_existence == []:
        content = re.sub('<table.*?>', '<table border="1" cellspacing="0" align="center">', content)
        if re.findall('</table>', content):
            pass
        else:
            content = content + '</table>'
    else:
        pass

    # 这一组是处理 li ul 标签，ul啥都不保留，
    content = re.sub('<ul.*?>', '<ul>', str(content))
    content = re.sub('<li.*?>', '<li style="list-style-type:none;">', str(content))

    # 这一组是去除<?xml:namespace prefix = o ns = "urn:schemas-microsoft-com:office:office" /><o:p></o:p>
    content = re.sub(r'<?xml:namespace .*?>', '', content, flags=re.I)
    content = re.sub(r'<o:p.*?>', '', content, flags=re.I)
    content = re.sub(r'</o:p>', '', content, flags=re.I)

    content = re.sub(r'<st1:chsdate .*?>', '', content, flags=re.I)
    content = re.sub(r'</st1:chsdate>', '', content, flags=re.I)
    content = re.sub('<script.*?>.*?</script>', '', content, flags=re.I | re.S)
    content = re.sub('<style.*?>.*?</style>', '', content, flags=re.I | re.S)

    # <v:line></v:line> 什么鬼的直接链接符啥？啥玩意儿这是，一脸懵逼
    content = re.sub(r'<v:line.*?>.*?</v:line>', '', content, flags=re.S | re.I)

    # 处理单引号问题，文中单引号往数据库插入数据是行不通的
    content = re.sub(r"'", '"', content, flags=re.I)

    # 因系统的需要，在块级元素后面加一个《br》
    content = content.replace("</div>", "</div><br/>")
    content = content.replace("</h1>", "</h1><br/>")
    content = content.replace("</h2>", "</h2><br/>")
    content = content.replace("</h3>", "</h3><br/>")
    content = content.replace("</h4>", "</h4><br/>")
    content = content.replace("</h5>", "</h5><br/>")
    content = content.replace("</h6>", "</h6><br/>")
    content = content.replace("</hr>", "</hr><br/>")
    content = content.replace("</p>", "</p><br/>")
    content = content.replace("</table>", "</table><br/>")

    return content


# 上传到数据库操作
def upload_to_database(sql):
    connect_cursor = 链接数据库.get_connect_cursor()
    connect = connect_cursor[0]
    cursor = connect_cursor[1]
    链接数据库.insert(cursor, sql)
    链接数据库.break_connect(connect)


# 查询操作
def selct_info():
    pass





# 处理附加名：将文中的附件名添加到数据表中
# 前提：先将附件全部处理好（缺失补充、没改链接的修改链接）
def check_adj_name(title, cont, save_path_father_server, save_path_local):
    all_a = re.findall(r'<a.*?href=".*?".*?>.*?</a>', str(cont), flags=re.I)
    the_server_adj_list = " "
    the_local_adjlist = ""
    if all_a != []:
        for a in all_a:
            # 循环所有的a标签,再遍历每一个a标签，在每一个a标签中去匹配是否是附件形式 ，不是就直接替换掉，是的话在进行下一步：匹配附件a标签中冲链接和文本能容
            annex_a = re.findall(
                r'<a.*?href=".*?(pdf|docx|doc|xlsx|xls|rar|zip|jpeg|jpg|png|gif|txt|7z|gz)".*?>.*?</a>', a)
            if annex_a:

                # 假如是附件的形式，使用all_a结果集中每个元素的第一个值（超链接）和第二个值（文本内容）反向匹配这个标签的全部内容，并生成新的本地超链接a标签
                rs_a = re.findall('<a.*?href="(.*?)".*?>(.*?)</a>', a)
                a_hyper = rs_a[0][0]    # a标签的超链接
                a_annex_name = a_hyper[a_hyper.rfind("/") + 1:]  # 附件名字

                # 匹配本地附件是否存在
                # 超链接地址已经做了修改是属于我们服务器上面的地址，文中的超链接修改过了，开始检查本地附件
                if a_hyper.find("/datafolder/") != -1:
                    the_server_adj_list += "{0}, ".format(save_path_father_server + str(a_hyper).replace("/", '\\'))
                    the_local_adjlist +=  "{0}, ".format(save_path_local + str(a_hyper).replace("/", '\\'))
                else:
                    raise Exception("抱歉，这篇文章标题：{0} 的这个 a 标签 {1} 在全文中没有进行相应的修改，请手动修改并检查该附件是否存在，然后再来运行本方法吧！！！！".format(title, a))
            else:
                raise Exception("这篇文章标题：{0} 的这个 a 标签{1} 不是附件形式，前面没有做出相应的修改，请手动修改吧！！！！".format(title, a))

    all_img = re.findall('<img.*?src=".*?".*?>', cont)
    if all_img != []:
        for img in all_img:
            picture = re.findall('<img.*?src=".*?(jpeg|jpg|png|gif)".*?>', img)
            if picture != []:
                rs_img = re.findall('<img.*?src="(.*?)".*?>', img)
                img_hyper = rs_img[0]  # 图片文中地址
                img_annex_name = img_hyper[img_hyper.rfind("/") + 1:]
                if img_hyper.find("/datafolder/") != -1:
                    the_server_adj_list += "{0}, ".format(save_path_father_server + str(img_hyper).replace("/", '\\'))
                    the_local_adjlist += "{0}, ".format(save_path_local + str(img_hyper).replace("/", '\\'))
                else:
                    raise Exception(
                        "抱歉，这篇文章标题：{0} 的这个 img 标签 {1} 在全文中没有进行相应的修改，请手动修改并检查该附件是否存在，然后再来运行本方法吧！！！！".format(title, img))
            else:
                raise Exception(
                    "抱歉，这篇文章标题：{0} 的这个 img 标签 {1} 不是附件形式，前面没有做出相应的修改，请手动修改吧！！！！".format(title, img))

    return the_server_adj_list, the_local_adjlist

# str1 = """
# <a href="/datafolder/知识产权局/知识产权工作/20190430213953632793.pdf">20190430213953632793.pdf</a>
# <a href="/datafolder/知识产权局/知识产权工作/20190430213953632794.pdf">20190430213953632794.pdf</a>
# <a href="/datafolder/知识产权局/知识产权工作/20190430213953632795.pdf">20190430213953632795.pdf</a>
#
# """
# ss = check_adj_name('ceshi', str1, 'E:')
# print(ss)